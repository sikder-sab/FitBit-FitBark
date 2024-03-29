import fitbit
import oauth2
import requests
import threading
from selenium import webdriver
from bs4 import BeautifulSoup
import time
import threading
import re
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime


def getDateStrings(dates):
    startDateString = dates[0].strftime('%Y-%m-%d')
    endDateString = dates[1].strftime('%Y-%m-%d')
    return startDateString, endDateString


filepath = "test logins.xlsx"
userDataWorkbook = openpyxl.load_workbook(filepath, data_only=True)
userDataSheet = userDataWorkbook.active
masterCSVString = ""

index = 0
for row in userDataSheet.iter_rows():
    # if index == 2:
    #     break

    rowValue = row[0].row
    participantType = userDataSheet.cell(row=rowValue, column=1).value
    uniqueID = userDataSheet.cell(row=rowValue, column=2).value
    familyID = userDataSheet.cell(row=rowValue, column=3).value
    fitbitEmail = userDataSheet.cell(row=rowValue, column=4).value
    fitbitPassword = userDataSheet.cell(row=rowValue, column=5).value
    studygroup = userDataSheet.cell(row=rowValue, column=8).value
    startDate = userDataSheet.cell(row=rowValue, column=9).value
    endDate = userDataSheet.cell(row=rowValue, column=10).value

    if participantType is not None:
        dates = getDateStrings([startDate,endDate])
        print(dates)

    