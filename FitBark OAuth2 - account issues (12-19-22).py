import fitbark
import requests
from selenium import webdriver
from bs4 import BeautifulSoup
import time
import threading
import re
from datetime import datetime, timedelta
import requests
import http.client
import json
import os
import openpyxl
from openpyxl import Workbook

class APIcalls():
    @staticmethod
    def getAccesstoken(client_id,client_secret,callback_uri,authencode,accesstoken_url):
        data = {'client_id': client_id,'client_secret':client_secret,'grant_type':'authorization_code','redirect_uri': callback_uri, 'code':authencode}
        try:
            access_token_response = requests.post(accesstoken_url, data=data)
            statusCode = access_token_response.status_code
            #print access_token_response.url, access_token_response.headers
            return access_token_response 
        except requests.ConnectionError:
            print("failed to connect")
            return False

    @staticmethod
    def dogActivityHTTPSRequest(data, accesstoken):
        data = data
        headers = { 'Authorization': "Bearer "+ accesstoken, 'Content-Type':'application/json'}
        try:
            access_token_response = requests.post("https://app.fitbark.com/api/v2/activity_series", headers=headers,data=data)
            return access_token_response 
        except requests.ConnectionError:
            print("failed to connect")
            return False

    @staticmethod
    def getActivityDataForDog(client_id, client_secret, accesstoken, dates):
        #GET DATA
        fitbarkprofile =  fitbark.FitBark(client_id, client_secret, access_token=accesstoken, refresh_token='<refresh_token>')
        profile_get = fitbarkprofile.user_profile_get()
        allRelatedDogsData = fitbarkprofile.user_related_dogs_get()
        dogsData = allRelatedDogsData["dog_relations"]

        slug = ""
        for dog in dogsData:
            status = dog["status"]
            if status == "OWNER":
                slug = dog["dog"]["slug"]
                break
        # print("ok",slug)

        activityData = APIcalls.getAllDataForUserByInterval(accesstoken, slug, dates)
        return activityData
        # todatestring = extractiondate.strftime('%Y-%m-%d')
        # fromdatestring = (extractiondate - timedelta(days=42)).strftime('%Y-%m-%d') #The longest we can go back in terms of data is 42 days. "The maximum range is 42 days with daily resolution. - https://www.fitbark.com/dev/ for GET ACTIVITY SERIES"

        # jsondata = {
        # "activity_series":{
        #   "to": todatestring,
        #   "resolution": "DAILY", 
        #   "from": fromdatestring, 
        #   "slug": slug
        #   }
        # }
        # jsonstring = json.dumps(jsondata)
        # activitydatabefore = APIcalls.dogActivityHTTPSRequest(jsonstring, ACCESS_TOKEN)
        # print(activitydatabefore)
        # activitydata = activitydatabefore.json()
        # return activitydata

    @staticmethod
    def getAllDataForUserByInterval(ACCESS_TOKEN, slug, dates):
        increment = 42
        startdateTimestamp = datetime.strptime(dates[0], '%Y-%m-%d')
        currentDate = datetime.strptime(dates[1], '%Y-%m-%d')
        activityJsonData = []
        print(startdateTimestamp, currentDate)

        while currentDate > startdateTimestamp:
            endTimestamp = startdateTimestamp + timedelta(days=increment)

            startString = datetime.strftime(startdateTimestamp, '%Y-%m-%d')
            endString = datetime.strftime(endTimestamp, '%Y-%m-%d')
            # print(startString,endString )

            jsondata = {
            "activity_series":{
                "to": endString,
                "resolution": "DAILY", 
                "from": startString, 
                "slug": slug
                }
            }
            jsonstring = json.dumps(jsondata)
            activitydatabefore = APIcalls.dogActivityHTTPSRequest(jsonstring, ACCESS_TOKEN)
            # print(activitydatabefore)
            activitydata = activitydatabefore.json()
            print(activitydata)
            parsedActivityData = activitydata["activity_series"]["records"]
            if len(parsedActivityData) > 0:
                activityJsonData += parsedActivityData
            # print(endString,"------data in the end---------------------------",parsedActivityData)

            startdateTimestamp = endTimestamp
        return activityJsonData

def getDateStrings(dates):
    startDateString = dates[0].strftime('%Y-%m-%d')
    endDateString = dates[1].strftime('%Y-%m-%d')
    return startDateString, endDateString

def parseActivityDataAsCSV(json,participantType,uniqueID,familyID,fitbitEmail,studygroup):
    activityList = json
    csv = ""

    for activity in activityList:
        date = activity["date"]
        min_rest = activity["min_rest"]
        min_play = activity["min_play"]
        min_active = activity["min_active"]
        activity_value = activity["activity_value"]

        csv = csv + date + "," + str(participantType) + "," + str(uniqueID) + "," + str(familyID) + "," + str(fitbitEmail) + "," + str(studygroup) + "," + str(min_rest) + "," + str(min_play) + "," + str(min_active) + "," + str(activity_value) 
        csv += "\n"
    return csv


def createNewFolderIfNeeded(newfolder):
    try:
        os.makedirs(newfolder)
    except OSError:
        pass


def checkURLAndGetCode(webDriver):
    currentURL = webDriver.current_url
    print(currentURL)
    if "code=" in currentURL:
        print("yes its in here")
        regexQuery = "(code)=(\\w*)&"
        regexSearch = re.search(regexQuery, currentURL)
        code = regexSearch.group(2)
        webDriver.close()
        return code
    elif "&scope" in currentURL:
        print("we are in authorize")
        time.sleep(3)
        allowAuthorize = webDriver.find_element_by_name("commit")
        allowAuthorize.click()
        return checkURLAndGetCode(webDriver)
    elif "login_attempt" in currentURL:
        webdriver.close()
        return "login issue"
    else:
        timer = threading.Timer(4.0, checkURLAndGetCode(webDriver))
        timer.start


def getDataForUserFromCredentials(credentialRow):
    rowValue = credentialRow[0].row
    participantType = userDataSheet.cell(row=rowValue, column=1).value
    uniqueID = userDataSheet.cell(row=rowValue, column=2).value
    familyID = userDataSheet.cell(row=rowValue, column=3).value
    fitbarkEmail = userDataSheet.cell(row=rowValue, column=6).value
    fitbarkPassword = userDataSheet.cell(row=rowValue, column=7).value
    studygroup = userDataSheet.cell(row=rowValue, column=8).value
    startDate = userDataSheet.cell(row=rowValue, column=9).value
    endDate = userDataSheet.cell(row=rowValue, column=10).value
    print(fitbarkEmail, fitbarkPassword)

    # print(participantType)
    if participantType is not None and fitbarkPassword is not None:

        client_id = "422c1260cf445d46b6b6e5dba9910dbabb9d706c4e5b64df4a71b86fa2fba351"
        client_secret = "f0d46a863cbfbd66f68b845e0c96343b2270b519b700de5f77cf54d272066260"
        callback_uri = "http://127.0.0.1:8080/"
        accesstoken_url = "https://app.fitbark.com/oauth/token"

        authd_client = fitbark.FitBarkOauth2Client(client_id, client_secret,access_token='<access_token>', refresh_token='<refresh_token>')
        tokenURL = authd_client.authorize_token_url(redirect_uri=callback_uri)
        authenticationUrl = str(tokenURL[0])
        print (authenticationUrl)

        options = webdriver.ChromeOptions()
        driver = webdriver.Chrome('/usr/local/bin/chromedriver', chrome_options=options) 
        driver.get(authenticationUrl)

        submitButton = driver.find_element_by_id("button-login")
        username = driver.find_element_by_id("username_or_email")
        password = driver.find_element_by_id("login_password")
        
        #start URL listener
        time.sleep(3)
        username.send_keys(fitbarkEmail)
        password.send_keys(fitbarkPassword)
        submitButton.click()

        authencode = checkURLAndGetCode(driver)
        if authencode is None:
            authencode = checkURLAndGetCode(driver)

        print("authencode", authencode)
        csvdata = fitbarkEmail + "," + authencode
        csvdata += "\n"
        return csvdata

def makeMasterCSV(arrayOfCSVs):
    masterCSV = ""
    for csv in arrayOfCSVs:
        masterCSV = masterCSV + csv
    return masterCSV


#-------------------------------------------- code starts here ---------------------------------------------------------------------------------------
filepath = "test logins.xlsx"
userDataWorkbook = openpyxl.load_workbook(filepath, data_only=True)
userDataSheet = userDataWorkbook.active

masterCSVString = ""
index = 0
for row in userDataSheet.iter_rows():
    userCSVData = getDataForUserFromCredentials(row)

    if userCSVData is not None:
        masterCSVString += userCSVData

    index += 1

header = "Email,Issue"
masterCSVString = header + "\n" + masterCSVString

print(masterCSVString)




today = datetime.today()
dateString = today.strftime("%m-%d-%Y")
newfolder = "./FitBark Issues/"+ dateString + "/"
createNewFolderIfNeeded(newfolder)
with open(newfolder + "FitBark - all users" + ".csv","w") as w:
    w.write(masterCSVString)


# dateString = extractiondate.strftime("%m-%d-%Y")
# newfolder = "./FitBark/"+ dateString + "/"
# createNewFolderIfNeeded(newfolder)
# with open(newfolder + fitbarkEmail.replace(' ', '') + "test.csv","w") as w:
#   w.write(csvdata)




