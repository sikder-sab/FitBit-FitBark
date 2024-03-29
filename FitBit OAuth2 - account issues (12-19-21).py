import fitbit
import oauth2
import requests
import threading
from selenium import webdriver
from bs4 import BeautifulSoup
import time
import threading
import re
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime

def createNewFolderIfNeeded(newfolder):
    try:
        os.makedirs(newfolder)
    except OSError:
        pass

def getDateStrings(dates):
    startDateString = dates[0].strftime('%Y-%m-%d')
    endDateString = dates[1].strftime('%Y-%m-%d')
    return startDateString, endDateString

class APIcalls():
    @staticmethod
    def getAccesstoken(authencode):
        data = {'client_id': client_id,'grant_type':'authorization_code','redirect_uri': callback_uri, 'code':authencode}
        try:
            access_token_response = requests.post(accesstoken_url, data=data)
            statusCode = access_token_response.status_code
            print (access_token_response.url, access_token_response.headers, statusCode)
            return access_token_response 
        except requests.ConnectionError:
            print("failed to connect")
            return False


def checkURLAndGetCode(webDriver):
    currentURL = webDriver.current_url
    print(currentURL)
    time.sleep(3)
    if "code=" in currentURL:
        print("yes its in here")
        regexQuery = "(code)=(\\w*)&"
        regexSearch = re.search(regexQuery, currentURL)
        code = regexSearch.group(2)
        print(code)
        return code
    elif "&scope=activity+nutrition+heartrate" in currentURL:
        print("we are in scopes")
        time.sleep(3)
        allowAllScopes = driver.find_element_by_id("selectAllScope")
        allowAllScopes.click()

        allowButton = driver.find_element_by_id("allow-button")
        allowButton.click()
        return checkURLAndGetCode(webDriver)
    elif "mfa" in currentURL:
        print("2 factor authen")
        return "2 factor authentication"
    else:
        try:
            loginissue = driver.find_element_by_id("ember589")
            return "login issue"
        except:
            timer = threading.Timer(3.0, checkURLAndGetCode(webDriver))
            timer.start


#----------------------------------- start code -------------------------------------------------------------------- 
filepath = "./test logins.xlsx"
userDataWorkbook = openpyxl.load_workbook(filepath, data_only=True)
userDataSheet = userDataWorkbook.active
masterCSVString = ""

index = 0
for row in userDataSheet.iter_rows():
    if index == 3:
        break

    rowValue = row[0].row
    participantType = userDataSheet.cell(row=rowValue, column=1).value
    uniqueID = userDataSheet.cell(row=rowValue, column=2).value
    familyID = userDataSheet.cell(row=rowValue, column=3).value
    fitbitEmail = userDataSheet.cell(row=rowValue, column=4).value
    fitbitPassword = userDataSheet.cell(row=rowValue, column=5).value
    studygroup = userDataSheet.cell(row=rowValue, column=8).value
    startDate = userDataSheet.cell(row=rowValue, column=9).value
    endDate = userDataSheet.cell(row=rowValue, column=10).value

    if participantType is not None:
        # fitbitData = getFitBitCSVData(participantType, uniqueID, familyID, fitbitEmail, fitbitPassword, studygroup)
        # print(fitbitData)

        client_id =  "22BVY2"
        client_secret = "1f42a3dd1c48dda6db142e3d5ae9999e"
        callback_uri = "http://127.0.0.1:8080/"
        accesstoken_url = "https://api.fitbit.com/oauth2/token"

        authd_client = fitbit.FitbitOauth2Client(client_id, client_secret,access_token='<access_token>', refresh_token='<refresh_token>')
        tokenURL = authd_client.authorize_token_url(redirect_uri=callback_uri)
        authenticationUrl = str(tokenURL[0])
        print (authenticationUrl)

        options = webdriver.ChromeOptions()
        driver = webdriver.Chrome('/usr/local/bin/chromedriver', chrome_options=options) 
        driver.get(authenticationUrl)

        #start URL listener
        time.sleep(3)
        username = driver.find_element_by_id("ember664")
        password = driver.find_element_by_id("ember665")
        submitButton = driver.find_element_by_id("ember705")

        username.send_keys(fitbitEmail)
        password.send_keys(fitbitPassword)
        submitButton.click()

        authencode = checkURLAndGetCode(driver)
        if authencode is None:
            authencode = checkURLAndGetCode(driver)
        driver.close()

        print("authencode", authencode)

        csvdata = str(fitbitEmail) + "," + str(authencode)
        csvdata += "\n"

        masterCSVString += csvdata
        index += 1

header = "Email" + "," + "Issue"
masterCSVString = header + "\n" + masterCSVString
print(masterCSVString)

today = datetime.today()
dateString = today.strftime("%m-%d-%Y")
newfolder = "./FitBit Issues/"+ dateString + "/"
createNewFolderIfNeeded(newfolder)
with open(newfolder + "FitBit - all users" + ".csv","w") as w:
    w.write(masterCSVString)
        


